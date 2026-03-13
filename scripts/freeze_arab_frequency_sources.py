"""Freeze Arab frequency sources and stage AVDB liftover outputs.

This script keeps the pipeline simple and explicit:
1. Freeze SHGP raw bytes to the GCS raw vault with source metadata.
2. Freeze the AVDB workbook to the GCS raw vault with workbook provenance.
3. Parse AVDB genomic HGVS strings on GRCh37.
4. Lift the parsed genomic interval to GRCh38 with the official Ensembl assembly map API.
5. Write a Parquet checkpoint plus a JSON report to GCS for the successful AVDB liftover stage.

The script does not normalize alleles. That remains a later T003 phase.
"""

from __future__ import annotations

import datetime as dt
import gzip
import hashlib
import json
import re
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

import pandas as pd
from google.cloud import storage
from openpyxl import load_workbook

ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ID: Final[str] = "genome-services-platform"
BUCKET_NAME: Final[str] = "mahmoud-arab-acmg-research-data"
SNAPSHOT_DATE: Final[str] = dt.date.today().isoformat()
RUN_ID: Final[str] = f"arab-frequency-freeze-{dt.datetime.now(dt.UTC).strftime('%Y%m%dT%H%M%SZ')}"
TIMESTAMP_UTC: Final[str] = dt.datetime.now(dt.UTC).isoformat()

SHGP_LOCAL_FILE: Final[Path] = Path("/Users/macbookpro/Desktop/storage/raw/shgp/Saudi_Arabian_Allele_Frequencies.txt")
AVDB_LOCAL_FILE: Final[Path] = Path("/Users/macbookpro/Desktop/storage/raw/uae/avdb_uae.xlsx")

SHGP_FIGSHARE_API: Final[str] = "https://api.figshare.com/v2/articles/28059686"
AVDB_DOWNLOADS_URL: Final[str] = "https://avdb-arabgenome.ae/downloads"
GRCH37_ASSEMBLY_REPORT_URL: Final[str] = (
    "https://ftp.ncbi.nlm.nih.gov/genomes/all/GCF/000/001/405/"
    "GCF_000001405.25_GRCh37.p13/GCF_000001405.25_GRCh37.p13_assembly_report.txt"
)
GRCH38_ASSEMBLY_REPORT_URL: Final[str] = (
    "https://ftp.ncbi.nlm.nih.gov/genomes/all/GCF/000/001/405/"
    "GCF_000001405.40_GRCh38.p14/GCF_000001405.40_GRCh38.p14_assembly_report.txt"
)
ENSEMBL_MAP_URL_TEMPLATE: Final[str] = (
    "https://grch37.rest.ensembl.org/map/human/GRCh37/"
    "{chrom}:{start}..{end}:1/GRCh38?content-type=application/json"
)

SHGP_SOURCE_KEY: Final[str] = "shgp_saudi_af"
SHGP_SOURCE_VERSION: Final[str] = "figshare-28059686-v1"
AVDB_SOURCE_KEY: Final[str] = "avdb_uae"

SHGP_GCS_PREFIX: Final[str] = (
    f"raw/sources/{SHGP_SOURCE_KEY}/version={SHGP_SOURCE_VERSION}/snapshot_date={SNAPSHOT_DATE}"
)
AVDB_WORKBOOK_CREATED: Final[str] = "2025-06-27T10:05:29"
AVDB_SOURCE_VERSION: Final[str] = f"workbook-created-{AVDB_WORKBOOK_CREATED[:10]}"
AVDB_GCS_PREFIX: Final[str] = (
    f"raw/sources/{AVDB_SOURCE_KEY}/version={AVDB_SOURCE_VERSION}/build=GRCh37/snapshot_date={SNAPSHOT_DATE}"
)
AVDB_LIFTOVER_PREFIX: Final[str] = (
    f"frozen/harmonized/source={AVDB_SOURCE_KEY}/version={AVDB_SOURCE_VERSION}/"
    f"stage=liftover/build=GRCh37_to_GRCh38/snapshot_date={SNAPSHOT_DATE}"
)

AVDB_SHEET_NAME: Final[str] = "Sheet1"
AVDB_PARQUET_FILE: Final[Path] = ROOT / "tmp_avdb_liftover.parquet"
AVDB_REPORT_FILE: Final[Path] = ROOT / "tmp_avdb_liftover_report.json"


@dataclass(frozen=True)
class ParsedHgvs:
    accession: str
    start37: int
    end37: int
    event_type: str
    source_ref: str | None
    source_alt: str | None
    inserted_sequence: str | None


SUBSTITUTION_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)([ACGT]+)>([ACGT]+)$"
)
DELETION_SINGLE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)del$"
)
DELETION_RANGE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)_([0-9]+)del$"
)
INSERTION_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)_([0-9]+)ins([ACGT]+)$"
)
DUPLICATION_SINGLE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)dup$"
)
DUPLICATION_RANGE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)_([0-9]+)dup$"
)
DELINS_SINGLE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)delins([ACGT]+)$"
)
DELINS_RANGE_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^(NC_[0-9]+\.[0-9]+):g\.([0-9]+)_([0-9]+)delins([ACGT]+)$"
)


# [AI-Agent: Codex]: IO helper 1 - curl is more reliable than Python SSL on this workstation,
# so all provider metadata and assembly-report fetches go through curl.
def curl_bytes(url: str) -> bytes:
    return subprocess.run(
        ["curl", "-sSL", url],
        check=True,
        capture_output=True,
        text=False,
    ).stdout


def curl_json(url: str) -> dict[str, Any]:
    return json.loads(curl_bytes(url))


def file_hash(path: Path, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def text_line_count(path: Path) -> int:
    count = 0
    with path.open("rb") as handle:
        for _ in handle:
            count += 1
    return count


def gzip_line_count(path: Path) -> int:
    count = 0
    with gzip.open(path, "rt", encoding="utf-8") as handle:
        for _ in handle:
            count += 1
    return count


def upload_file(storage_client: storage.Client, local_path: Path, object_name: str, content_type: str | None = None) -> str:
    bucket = storage_client.bucket(BUCKET_NAME)
    blob = bucket.blob(object_name)
    blob.upload_from_filename(str(local_path), content_type=content_type)
    return f"gs://{BUCKET_NAME}/{object_name}"


def upload_text(storage_client: storage.Client, object_name: str, text: str, content_type: str) -> str:
    bucket = storage_client.bucket(BUCKET_NAME)
    blob = bucket.blob(object_name)
    blob.upload_from_string(text, content_type=content_type)
    return f"gs://{BUCKET_NAME}/{object_name}"


def fetch_shgp_file_metadata() -> dict[str, Any]:
    article = curl_json(SHGP_FIGSHARE_API)
    for file_entry in article["files"]:
        if file_entry["name"] == SHGP_LOCAL_FILE.name:
            return {
                "article_title": article["title"],
                "doi": article["doi"],
                "published_date": article["published_date"],
                "modified_date": article["modified_date"],
                "upstream_file_url": file_entry["download_url"],
                "upstream_md5": file_entry["supplied_md5"],
                "upstream_size": int(file_entry["size"]),
            }
    raise RuntimeError(f"Could not find {SHGP_LOCAL_FILE.name} in the SHGP Figshare article metadata.")


def load_accession_to_chrom_map() -> dict[str, str]:
    report_text = curl_bytes(GRCH37_ASSEMBLY_REPORT_URL).decode("utf-8")
    mapping: dict[str, str] = {}
    for line in report_text.splitlines():
        if not line or line.startswith("#"):
            continue
        columns = line.split("\t")
        if columns[1] != "assembled-molecule":
            continue
        chrom = columns[0]
        refseq_accession = columns[6]
        mapping[refseq_accession] = chrom
    return mapping


def parse_hgvs_genomic37(value: str | None) -> ParsedHgvs | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    match = SUBSTITUTION_PATTERN.match(text)
    if match:
        return ParsedHgvs(
            accession=match.group(1),
            start37=int(match.group(2)),
            end37=int(match.group(2)),
            event_type="substitution",
            source_ref=match.group(3),
            source_alt=match.group(4),
            inserted_sequence=None,
        )

    match = DELETION_SINGLE_PATTERN.match(text)
    if match:
        return ParsedHgvs(match.group(1), int(match.group(2)), int(match.group(2)), "deletion", None, None, None)

    match = DELETION_RANGE_PATTERN.match(text)
    if match:
        return ParsedHgvs(match.group(1), int(match.group(2)), int(match.group(3)), "deletion", None, None, None)

    match = INSERTION_PATTERN.match(text)
    if match:
        return ParsedHgvs(
            accession=match.group(1),
            start37=int(match.group(2)),
            end37=int(match.group(3)),
            event_type="insertion",
            source_ref=None,
            source_alt=None,
            inserted_sequence=match.group(4),
        )

    match = DUPLICATION_SINGLE_PATTERN.match(text)
    if match:
        return ParsedHgvs(match.group(1), int(match.group(2)), int(match.group(2)), "duplication", None, None, None)

    match = DUPLICATION_RANGE_PATTERN.match(text)
    if match:
        return ParsedHgvs(match.group(1), int(match.group(2)), int(match.group(3)), "duplication", None, None, None)

    match = DELINS_SINGLE_PATTERN.match(text)
    if match:
        return ParsedHgvs(
            accession=match.group(1),
            start37=int(match.group(2)),
            end37=int(match.group(2)),
            event_type="delins",
            source_ref=None,
            source_alt=None,
            inserted_sequence=match.group(3),
        )

    match = DELINS_RANGE_PATTERN.match(text)
    if match:
        return ParsedHgvs(
            accession=match.group(1),
            start37=int(match.group(2)),
            end37=int(match.group(3)),
            event_type="delins",
            source_ref=None,
            source_alt=None,
            inserted_sequence=match.group(4),
        )

    return None


def ensembl_liftover(chrom: str, start37: int, end37: int, cache: dict[tuple[str, int, int], dict[str, Any]]) -> dict[str, Any]:
    cache_key = (chrom, start37, end37)
    if cache_key in cache:
        return cache[cache_key]

    url = ENSEMBL_MAP_URL_TEMPLATE.format(chrom=chrom, start=start37, end=end37)
    last_error: str | None = None
    for attempt in range(3):
        result = subprocess.run(["curl", "-sS", "--fail", url], capture_output=True, text=True)
        if result.returncode == 0:
            payload = json.loads(result.stdout)
            mappings = payload.get("mappings", [])
            if len(mappings) != 1:
                mapped = {
                    "status": "failed",
                    "reason": f"expected_single_mapping_got_{len(mappings)}",
                    "chrom38": None,
                    "start38": None,
                    "end38": None,
                }
            else:
                mapped_record = mappings[0]["mapped"]
                mapped_length = int(mapped_record["end"]) - int(mapped_record["start"])
                input_length = end37 - start37
                if mapped_length != input_length:
                    mapped = {
                        "status": "failed",
                        "reason": f"length_changed_{input_length}_to_{mapped_length}",
                        "chrom38": None,
                        "start38": None,
                        "end38": None,
                    }
                else:
                    mapped = {
                        "status": "success",
                        "reason": "ok",
                        "chrom38": str(mapped_record["seq_region_name"]),
                        "start38": int(mapped_record["start"]),
                        "end38": int(mapped_record["end"]),
                    }
            cache[cache_key] = mapped
            return mapped

        last_error = result.stderr.strip() or f"curl_exit_{result.returncode}"
        time.sleep(0.5 * (attempt + 1))

    mapped = {
        "status": "failed",
        "reason": f"curl_error:{last_error or 'unknown'}",
        "chrom38": None,
        "start38": None,
        "end38": None,
    }
    cache[cache_key] = mapped
    return mapped


def freeze_shgp(storage_client: storage.Client) -> dict[str, Any]:
    metadata = fetch_shgp_file_metadata()
    sha256 = file_hash(SHGP_LOCAL_FILE, "sha256")
    md5 = file_hash(SHGP_LOCAL_FILE, "md5")
    row_count = text_line_count(SHGP_LOCAL_FILE) - 1

    if md5 != metadata["upstream_md5"]:
        raise RuntimeError(
            f"SHGP local MD5 mismatch. local={md5} upstream={metadata['upstream_md5']}"
        )

    raw_object = f"{SHGP_GCS_PREFIX}/{SHGP_LOCAL_FILE.name}"
    raw_uri = upload_file(storage_client, SHGP_LOCAL_FILE, raw_object, content_type="text/tab-separated-values")

    manifest = {
        "source": SHGP_SOURCE_KEY,
        "source_version": SHGP_SOURCE_VERSION,
        "snapshot_date": SNAPSHOT_DATE,
        "upstream_url": metadata["upstream_file_url"],
        "upstream_article_title": metadata["article_title"],
        "doi": metadata["doi"],
        "published_date": metadata["published_date"],
        "modified_date": metadata["modified_date"],
        "local_sha256": sha256,
        "local_md5": md5,
        "upstream_md5": metadata["upstream_md5"],
        "gcs_uri": raw_uri,
        "row_count": row_count,
        "notes": (
            "Local file matched the official Figshare MD5 exactly. "
            "The table is genome-wide, GRCh38, and exposes CHROM/POS/REF/ALT plus AC/AN."
        ),
    }
    manifest_uri = upload_text(
        storage_client,
        f"{SHGP_GCS_PREFIX}/manifest.json",
        json.dumps(manifest, indent=2, ensure_ascii=True) + "\n",
        content_type="application/json",
    )

    return {"raw_uri": raw_uri, "manifest_uri": manifest_uri, **manifest}


def avdb_workbook_metadata() -> dict[str, str]:
    workbook = load_workbook(AVDB_LOCAL_FILE, read_only=True, data_only=True)
    props = workbook.properties
    created = props.created.isoformat() if props.created else AVDB_WORKBOOK_CREATED
    modified = props.modified.isoformat() if props.modified else created
    creator = props.creator or "unknown"
    return {"created": created, "modified": modified, "creator": creator}


def freeze_avdb_raw(storage_client: storage.Client) -> dict[str, Any]:
    props = avdb_workbook_metadata()
    sha256 = file_hash(AVDB_LOCAL_FILE, "sha256")
    raw_object = f"{AVDB_GCS_PREFIX}/{AVDB_LOCAL_FILE.name}"
    raw_uri = upload_file(
        storage_client,
        AVDB_LOCAL_FILE,
        raw_object,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    workbook = load_workbook(AVDB_LOCAL_FILE, read_only=True, data_only=True)
    sheet = workbook[AVDB_SHEET_NAME]
    row_count = sheet.max_row - 1
    manifest = {
        "source": AVDB_SOURCE_KEY,
        "source_version": AVDB_SOURCE_VERSION,
        "snapshot_date": SNAPSHOT_DATE,
        "upstream_url": AVDB_DOWNLOADS_URL,
        "local_sha256": sha256,
        "gcs_uri": raw_uri,
        "row_count": row_count,
        "workbook_created": props["created"],
        "workbook_modified": props["modified"],
        "workbook_creator": props["creator"],
        "notes": (
            "Workbook is user-supplied from the official AVDB download surface. "
            "The public downloads page was verified, but a standalone public file release date "
            "was not recoverable from the provider page during this run."
        ),
    }
    manifest_uri = upload_text(
        storage_client,
        f"{AVDB_GCS_PREFIX}/manifest.json",
        json.dumps(manifest, indent=2, ensure_ascii=True) + "\n",
        content_type="application/json",
    )
    return {"raw_uri": raw_uri, "manifest_uri": manifest_uri, **manifest}


def build_avdb_liftover_frame(avdb_raw_uri: str, avdb_raw_sha256: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    accession_to_chrom = load_accession_to_chrom_map()
    frame = pd.read_excel(AVDB_LOCAL_FILE, sheet_name=AVDB_SHEET_NAME)
    frame = frame.rename(columns={"AlleleFrequencies ": "AlleleFrequencies"})

    liftover_cache: dict[tuple[str, int, int], dict[str, Any]] = {}
    records: list[dict[str, Any]] = []

    for offset, row in enumerate(frame.to_dict(orient="records"), start=2):
        parsed = parse_hgvs_genomic37(row.get("HGVS_Genomic_GRCh37"))
        if parsed is None:
            records.append(
                {
                    "source": AVDB_SOURCE_KEY,
                    "source_version": AVDB_SOURCE_VERSION,
                    "snapshot_date": SNAPSHOT_DATE,
                    "gene_symbol": row.get("Gene Symbol"),
                    "hgvs_c_clinically_relevant": row.get("HGVS c. (Clinically Relevant)"),
                    "hgvs_genomic_grch37": row.get("HGVS_Genomic_GRCh37"),
                    "allele_counts": row.get("AlleleCounts"),
                    "allele_frequency": row.get("AlleleFrequencies"),
                    "carrier_rate_2pq": row.get("Carrier rate (2pq)"),
                    "first_cousin_at_risk_rate": row.get("At-risk couples assuming first cousin marriages (2pq X1/8)"),
                    "source_artifact_uri": avdb_raw_uri,
                    "source_artifact_sha256": avdb_raw_sha256,
                    "source_sheet_name": AVDB_SHEET_NAME,
                    "source_row_number": offset,
                    "source_record_locator": f"sheet={AVDB_SHEET_NAME};row={offset}",
                    "parse_status": "missing_coordinates",
                    "source_build": "GRCh37",
                    "chrom37": None,
                    "start37": None,
                    "end37": None,
                    "event_type": None,
                    "source_ref": None,
                    "source_alt": None,
                    "inserted_sequence": None,
                    "liftover_status": "not_applicable",
                    "liftover_tool": "Ensembl REST map endpoint",
                    "liftover_tool_version": "grch37.rest.ensembl.org reviewed 2026-03-13",
                    "liftover_notes": "HGVS_Genomic_GRCh37 was empty or unsupported.",
                    "chrom38": None,
                    "start38": None,
                    "end38": None,
                    "norm_status": "not_applicable",
                    "norm_tool": None,
                    "norm_tool_version": None,
                    "norm_notes": "Liftover stage only; allele normalization deferred to Phase 3.",
                    "transform_run_id": RUN_ID,
                    "transform_timestamp_utc": TIMESTAMP_UTC,
                }
            )
            continue

        chrom37 = accession_to_chrom.get(parsed.accession)
        if chrom37 is None:
            liftover_payload = {
                "status": "failed",
                "reason": f"accession_not_in_assembly_report:{parsed.accession}",
                "chrom38": None,
                "start38": None,
                "end38": None,
            }
        else:
            liftover_payload = ensembl_liftover(chrom37, parsed.start37, parsed.end37, liftover_cache)

        records.append(
            {
                "source": AVDB_SOURCE_KEY,
                "source_version": AVDB_SOURCE_VERSION,
                "snapshot_date": SNAPSHOT_DATE,
                "gene_symbol": row.get("Gene Symbol"),
                "hgvs_c_clinically_relevant": row.get("HGVS c. (Clinically Relevant)"),
                "hgvs_genomic_grch37": row.get("HGVS_Genomic_GRCh37"),
                "allele_counts": row.get("AlleleCounts"),
                "allele_frequency": row.get("AlleleFrequencies"),
                "carrier_rate_2pq": row.get("Carrier rate (2pq)"),
                "first_cousin_at_risk_rate": row.get("At-risk couples assuming first cousin marriages (2pq X1/8)"),
                "source_artifact_uri": avdb_raw_uri,
                "source_artifact_sha256": avdb_raw_sha256,
                "source_sheet_name": AVDB_SHEET_NAME,
                "source_row_number": offset,
                "source_record_locator": f"sheet={AVDB_SHEET_NAME};row={offset}",
                "parse_status": "parsed",
                "source_build": "GRCh37",
                "chrom37": chrom37,
                "start37": parsed.start37,
                "end37": parsed.end37,
                "event_type": parsed.event_type,
                "source_ref": parsed.source_ref,
                "source_alt": parsed.source_alt,
                "inserted_sequence": parsed.inserted_sequence,
                "liftover_status": liftover_payload["status"],
                "liftover_tool": "Ensembl REST map endpoint",
                "liftover_tool_version": "grch37.rest.ensembl.org reviewed 2026-03-13",
                "liftover_notes": liftover_payload["reason"],
                "chrom38": liftover_payload["chrom38"],
                "start38": liftover_payload["start38"],
                "end38": liftover_payload["end38"],
                "norm_status": "not_applicable",
                "norm_tool": None,
                "norm_tool_version": None,
                "norm_notes": "Liftover stage only; allele normalization deferred to Phase 3.",
                "transform_run_id": RUN_ID,
                "transform_timestamp_utc": TIMESTAMP_UTC,
            }
        )

    liftover_frame = pd.DataFrame.from_records(records)
    report = build_avdb_liftover_report(liftover_frame)
    return liftover_frame, report


def build_avdb_liftover_report(frame: pd.DataFrame) -> dict[str, Any]:
    total_rows = int(len(frame))
    parse_success_rows = int((frame["parse_status"] == "parsed").sum())
    liftover_success_rows = int((frame["liftover_status"] == "success").sum())
    brca_rows = int(frame["gene_symbol"].isin(["BRCA1", "BRCA2"]).sum())
    brca_liftover_rows = int(
        frame.loc[frame["gene_symbol"].isin(["BRCA1", "BRCA2"]), "liftover_status"].eq("success").sum()
    )

    failure_examples = (
        frame.loc[frame["liftover_status"].ne("success"), [
            "source_row_number",
            "gene_symbol",
            "hgvs_genomic_grch37",
            "parse_status",
            "liftover_status",
            "liftover_notes",
        ]]
        .head(10)
        .to_dict(orient="records")
    )

    event_type_counts = (
        frame["event_type"].fillna("missing").value_counts().sort_index().to_dict()
    )

    return {
        "source": AVDB_SOURCE_KEY,
        "source_version": AVDB_SOURCE_VERSION,
        "snapshot_date": SNAPSHOT_DATE,
        "transform_run_id": RUN_ID,
        "transform_timestamp_utc": TIMESTAMP_UTC,
        "workflow_summary": (
            "Parsed AVDB HGVS_Genomic_GRCh37 rows, mapped RefSeq accessions to GRCh37 chromosomes via the "
            "official NCBI GRCh37 assembly report, then lifted genomic intervals to GRCh38 with the official "
            "Ensembl GRCh37->GRCh38 assembly map endpoint. Allele normalization was not performed in this stage."
        ),
        "official_sources": {
            "ncbi_grch37_assembly_report": GRCH37_ASSEMBLY_REPORT_URL,
            "ncbi_grch38_assembly_report": GRCH38_ASSEMBLY_REPORT_URL,
            "ensembl_map_api_template": ENSEMBL_MAP_URL_TEMPLATE,
            "avdb_downloads_page": AVDB_DOWNLOADS_URL,
        },
        "counts": {
            "total_rows": total_rows,
            "parse_success_rows": parse_success_rows,
            "parse_failure_rows": total_rows - parse_success_rows,
            "liftover_success_rows": liftover_success_rows,
            "liftover_failure_rows": total_rows - liftover_success_rows,
            "brca_rows": brca_rows,
            "brca_liftover_success_rows": brca_liftover_rows,
        },
        "event_type_counts": event_type_counts,
        "failure_examples": failure_examples,
        "use_decision": {
            "label": "reference_only",
            "reason": (
                "AVDB is scientifically useful and now coordinate-lifted, but it is still a small curated workbook "
                "rather than a broad population-frequency baseline. It should remain secondary/reference evidence "
                "unless a later normalization and overlap review justify stronger use."
            ),
        },
    }


def upload_avdb_liftover_artifacts(
    storage_client: storage.Client,
    liftover_frame: pd.DataFrame,
    report: dict[str, Any],
) -> dict[str, str]:
    liftover_frame.to_parquet(AVDB_PARQUET_FILE, index=False)
    AVDB_REPORT_FILE.write_text(json.dumps(report, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")

    parquet_uri = upload_file(
        storage_client,
        AVDB_PARQUET_FILE,
        f"{AVDB_LIFTOVER_PREFIX}/avdb_uae_liftover.parquet",
        content_type="application/octet-stream",
    )
    report_uri = upload_file(
        storage_client,
        AVDB_REPORT_FILE,
        f"{AVDB_LIFTOVER_PREFIX}/avdb_uae_liftover_report.json",
        content_type="application/json",
    )
    return {"parquet_uri": parquet_uri, "report_uri": report_uri}


def main() -> None:
    storage_client = storage.Client(project=PROJECT_ID)

    print("--- [Stage 1]: Freeze SHGP raw source ---")
    shgp_result = freeze_shgp(storage_client)
    print(
        f"SHGP raw_uri={shgp_result['raw_uri']} row_count={shgp_result['row_count']} "
        f"md5_match={shgp_result['local_md5'] == shgp_result['upstream_md5']}"
    )

    print("--- [Stage 2]: Freeze AVDB raw workbook ---")
    avdb_raw_result = freeze_avdb_raw(storage_client)
    print(f"AVDB raw_uri={avdb_raw_result['raw_uri']} row_count={avdb_raw_result['row_count']}")

    print("--- [Stage 3]: Build AVDB GRCh37 -> GRCh38 liftover artifact ---")
    liftover_frame, report = build_avdb_liftover_frame(
        avdb_raw_uri=avdb_raw_result["raw_uri"],
        avdb_raw_sha256=avdb_raw_result["local_sha256"],
    )
    artifact_uris = upload_avdb_liftover_artifacts(storage_client, liftover_frame, report)
    print(
        f"AVDB liftover parquet={artifact_uris['parquet_uri']} "
        f"success_rows={report['counts']['liftover_success_rows']} "
        f"failed_rows={report['counts']['liftover_failure_rows']}"
    )

    print("done")


if __name__ == "__main__":
    main()
