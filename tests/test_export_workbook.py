from io import BytesIO

from openpyxl import load_workbook

from ui.export_workbook import build_pre_gme_workbook_bytes, export_column_payload, export_header_columns, export_metadata_lines


def test_build_pre_gme_workbook_bytes_writes_metadata_header_and_rows():
    workbook_bytes = build_pre_gme_workbook_bytes(
        [
            {
                "CHROM": "chr13",
                "POS": 32315086,
                "END": 32315086,
                "ID": "chr13:32315086:A:G",
                "REF": "A",
                "ALT": "G",
                "VARTYPE": "SNV",
                "Repeat": None,
                "Segdup": None,
                "Blacklist": None,
                "GENE": "BRCA2",
                "EFFECT": None,
                "HGVS_C": None,
                "HGVS_P": None,
                "PHENOTYPES_OMIM": None,
                "PHENOTYPES_OMIM_ID": None,
                "INHERITANCE_PATTERN": None,
                "ALLELEID": "123",
                "CLNSIG": "Pathogenic",
                "TOPMED_AF": None,
                "TOPMed_Hom": None,
                "ALFA_AF": None,
                "ALFA_Hom": None,
                "GNOMAD_ALL_AF": 0.004,
                "gnomAD_all_Hom": 0,
                "GNOMAD_MID_AF": 0.0,
                "gnomAD_mid_Hom": 0,
                "ONEKGP_AF": None,
                "REGENERON_AF": None,
                "TGP_AF": None,
                "QATARI": None,
                "JGP_AF": None,
                "JGP_MAF": None,
                "JGP_Hom": None,
                "JGP_Het": None,
                "JGP_AC_Hemi": None,
                "SIFT_PRED": None,
                "POLYPHEN2_HDIV_PRED": None,
                "POLYPHEN2_HVAR_PRED": None,
                "PROVEAN_PRE": None,
                "CLNREVSTAT": "reviewed_by_expert_panel",
                "GNOMAD_GENOMES_AC": None,
                "GNOMAD_GENOMES_AN": None,
                "GNOMAD_GENOMES_AF": None,
                "GNOMAD_GENOMES_HOM": None,
                "GNOMAD_GENOMES_AF_AFR": None,
                "GNOMAD_GENOMES_AF_EUR_PROXY": None,
                "GNOMAD_EXOMES_AC": 4,
                "GNOMAD_EXOMES_AN": 1000,
                "GNOMAD_EXOMES_AF": 0.004,
                "GNOMAD_EXOMES_HOM": 0,
                "GNOMAD_EXOMES_AF_AFR": 0.0,
                "GNOMAD_EXOMES_AF_EUR_PROXY": 0.004,
                "GNOMAD_GENOMES_DEPTH": None,
                "GNOMAD_EXOMES_DEPTH": 88,
                "SOURCE_COUNT": 2,
                "PIPELINE_STAGE": "PRE_GME_REVIEW",
            }
        ],
        created_at="11/03/2026 12:00",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][0] == "ARAB_ACMG_PRE_GME_PIPELINE - Created on: 11/03/2026 12:00"
    assert rows[1][0].startswith("##WORKFLOW=")
    assert rows[len(export_metadata_lines("11/03/2026 12:00"))][0] == export_header_columns()[0]
    assert rows[-1][0] == "chr13"
    assert rows[-1][10] == "BRCA2"


def test_export_column_payload_marks_required_and_extra_columns():
    payload = export_column_payload()

    assert payload[0]["name"] == "CHROM"
    assert payload[0]["kind"] == "required"
    assert any(column["name"] == "PIPELINE_STAGE" and column["kind"] == "extra" for column in payload)
