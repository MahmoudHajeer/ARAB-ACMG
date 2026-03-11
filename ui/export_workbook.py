from __future__ import annotations

import datetime as dt
from io import BytesIO
from typing import Final, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

PRE_GME_EXPORT_FILENAME: Final[str] = "supervisor_variant_registry_brca_pre_gme_v1.xlsx"
PRE_GME_EXPORT_SHEET: Final[str] = "Pre_GME_Review"
PRE_GME_PIPELINE_STAGE: Final[str] = "PRE_GME_REVIEW"

PRE_GME_EXPORT_COLUMNS: Final[tuple[tuple[str, str], ...]] = (
    ("CHROM", "Canonical GRCh38 chromosome label."),
    ("POS", "1-based GRCh38 start position."),
    ("END", "1-based GRCh38 end position derived as POS + len(REF) - 1."),
    ("ID", "Stable export identifier. Mirrors the canonical variant key for reproducibility."),
    ("VARIANT_KEY", "Canonical cross-source variant key built as chrom:pos:ref:alt."),
    ("REF", "Canonical reference allele."),
    ("ALT", "Canonical alternate allele."),
    ("GENE", "Target BRCA gene assigned from the frozen Ensembl-backed gene window."),
    ("PRESENT_IN_CLINVAR", "True when the allele is present in the harmonized ClinVar BRCA table."),
    ("PRESENT_IN_GNOMAD_GENOMES", "True when the allele is present in the harmonized gnomAD genomes BRCA table."),
    ("PRESENT_IN_GNOMAD_EXOMES", "True when the allele is present in the harmonized gnomAD exomes BRCA table."),
    ("CLINVAR_IDS", "Distinct ClinVar identifiers observed for this allele."),
    ("CLNSIG", "Distinct ClinVar clinical-significance labels observed for this allele."),
    ("CLNREVSTAT", "Distinct ClinVar review-status labels observed for this allele."),
    ("CLINVAR_RECORD_COUNT", "How many ClinVar harmonized rows collapsed into this allele."),
    ("CLINVAR_GENEINFO_MATCH_COUNT", "How many ClinVar source rows agreed with the target BRCA gene in GENEINFO."),
    ("CLINVAR_GENEINFO_MISMATCH_COUNT", "How many ClinVar source rows within the window lacked a matching GENEINFO label."),
    ("GNOMAD_GENOMES_AC", "gnomAD genomes allele count."),
    ("GNOMAD_GENOMES_AN", "gnomAD genomes allele number."),
    ("GNOMAD_GENOMES_AF", "gnomAD genomes allele frequency."),
    ("GNOMAD_GENOMES_GRPMAX", "gnomAD genomes grpmax population label."),
    ("GNOMAD_GENOMES_GRPMAX_FAF95", "gnomAD genomes grpmax faf95 value."),
    ("GNOMAD_GENOMES_DEPTH", "gnomAD genomes depth slot from staging when available."),
    ("GNOMAD_GENOMES_AC_AFR", "gnomAD genomes African-ancestry allele count."),
    ("GNOMAD_GENOMES_AF_AFR", "gnomAD genomes African-ancestry allele frequency."),
    ("GNOMAD_GENOMES_AC_EUR_PROXY", "gnomAD genomes Europe proxy allele count."),
    ("GNOMAD_GENOMES_AF_EUR_PROXY", "gnomAD genomes Europe proxy allele frequency."),
    ("GNOMAD_EXOMES_AC", "gnomAD exomes allele count."),
    ("GNOMAD_EXOMES_AN", "gnomAD exomes allele number."),
    ("GNOMAD_EXOMES_AF", "gnomAD exomes allele frequency."),
    ("GNOMAD_EXOMES_GRPMAX", "gnomAD exomes grpmax population label."),
    ("GNOMAD_EXOMES_GRPMAX_FAF95", "gnomAD exomes grpmax faf95 value."),
    ("GNOMAD_EXOMES_DEPTH", "gnomAD exomes depth slot from staging when available."),
    ("GNOMAD_EXOMES_AC_AFR", "gnomAD exomes African-ancestry allele count."),
    ("GNOMAD_EXOMES_AF_AFR", "gnomAD exomes African-ancestry allele frequency."),
    ("GNOMAD_EXOMES_AC_EUR_PROXY", "gnomAD exomes Europe proxy allele count."),
    ("GNOMAD_EXOMES_AF_EUR_PROXY", "gnomAD exomes Europe proxy allele frequency."),
    ("SOURCE_COUNT", "Number of supporting non-GME source streams for the allele in this review stage."),
    ("PIPELINE_STAGE", "Pipeline checkpoint label. Fixed to PRE_GME_REVIEW for this export."),
)


def export_metadata_lines(created_at: str | None = None) -> list[str]:
    timestamp = created_at or dt.datetime.now(dt.UTC).strftime("%d/%m/%Y %H:%M")
    lines = [
        f"ARAB_ACMG_PRE_GME_PIPELINE - Created on: {timestamp}",
        '##WORKFLOW=<ID=PRE_GME_REVIEW,Description="BRCA review artifact generated after ClinVar + gnomAD harmonization and before adding GME frequencies">',
        '##SOURCE=<ID=CLINVAR,Description="Harmonized BRCA ClinVar source table in arab_acmg_harmonized">',
        '##SOURCE=<ID=GNOMAD_GENOMES,Description="Harmonized BRCA gnomAD genomes source table in arab_acmg_harmonized">',
        '##SOURCE=<ID=GNOMAD_EXOMES,Description="Harmonized BRCA gnomAD exomes source table in arab_acmg_harmonized">',
        '##RULE=<ID=WINDOWS,Description="BRCA1 and BRCA2 windows are frozen from Ensembl GRCh38 gene summaries">',
        '##RULE=<ID=PRE_GME,Description="This artifact intentionally excludes GME so the supervisor can review the global+clinical integration before Arab-specific frequency addition">',
    ]
    for column_name, description in PRE_GME_EXPORT_COLUMNS:
        lines.append(
            f'##INFO=<ID={column_name},Number=1,Type=String,Description="{description}">'
        )
    return lines


def export_header_columns() -> list[str]:
    return [column_name for column_name, _ in PRE_GME_EXPORT_COLUMNS]


def normalize_cell_value(value: object) -> object:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return value


def map_pre_gme_export_row(row: dict[str, object]) -> list[object]:
    return [
        row.get("chrom"),
        row.get("pos"),
        row.get("end_pos"),
        row.get("export_id"),
        row.get("variant_key"),
        row.get("ref"),
        row.get("alt"),
        row.get("gene_symbol"),
        normalize_cell_value(row.get("present_in_clinvar")),
        normalize_cell_value(row.get("present_in_gnomad_genomes")),
        normalize_cell_value(row.get("present_in_gnomad_exomes")),
        row.get("clinvar_ids"),
        row.get("clinvar_significance_values"),
        row.get("clinvar_review_status_values"),
        row.get("clinvar_record_count"),
        row.get("clinvar_gene_info_match_count"),
        row.get("clinvar_gene_info_mismatch_count"),
        row.get("gnomad_genomes_ac"),
        row.get("gnomad_genomes_an"),
        row.get("gnomad_genomes_af"),
        row.get("gnomad_genomes_grpmax"),
        row.get("gnomad_genomes_grpmax_faf95"),
        row.get("gnomad_genomes_depth"),
        row.get("gnomad_genomes_ac_afr"),
        row.get("gnomad_genomes_af_afr"),
        row.get("gnomad_genomes_ac_eur_proxy"),
        row.get("gnomad_genomes_af_eur_proxy"),
        row.get("gnomad_exomes_ac"),
        row.get("gnomad_exomes_an"),
        row.get("gnomad_exomes_af"),
        row.get("gnomad_exomes_grpmax"),
        row.get("gnomad_exomes_grpmax_faf95"),
        row.get("gnomad_exomes_depth"),
        row.get("gnomad_exomes_ac_afr"),
        row.get("gnomad_exomes_af_afr"),
        row.get("gnomad_exomes_ac_eur_proxy"),
        row.get("gnomad_exomes_af_eur_proxy"),
        row.get("source_count"),
        PRE_GME_PIPELINE_STAGE,
    ]


def build_pre_gme_workbook_bytes(
    rows: Iterable[dict[str, object]],
    created_at: str | None = None,
) -> bytes:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(PRE_GME_EXPORT_SHEET)

    title_font = Font(bold=True, size=12, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1D4E89")
    info_font = Font(color="0B304F")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="163B64")

    for index, line in enumerate(export_metadata_lines(created_at=created_at), start=1):
        cell = WriteOnlyCell(worksheet, value=line)
        cell.alignment = Alignment(wrap_text=True)
        if index == 1:
            cell.font = title_font
            cell.fill = title_fill
        else:
            cell.font = info_font
        worksheet.append([cell])

    header_row: list[WriteOnlyCell] = []
    for header in export_header_columns():
        cell = WriteOnlyCell(worksheet, value=header)
        cell.font = header_font
        cell.fill = header_fill
        header_row.append(cell)
    worksheet.append(header_row)

    for row in rows:
        worksheet.append(map_pre_gme_export_row(row))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
